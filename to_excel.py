import openpyxl
from xlutils.copy import copy
import xlwt
import xlrd
import os

def train_loss2excel(step, cross_entropy, Attention_name, dir):
    #path = os.path.join('./Result/ResNet/cifar10_ResNet18_loss_train.xlsx')
    path = os.path.join(dir)
    p = os.path.exists(path)
    loss = cross_entropy
    st = step
    if p:

        rb = openpyxl.load_workbook(path)
        wb = rb.sheetnames[0]
        # wb = copy(rb)
        s = rb[wb]
        #s = rb.get_sheet_by_name(wb[0])
        # s = wb.get_sheet(0)

        for n in range(1, st + 1):
            # s.write(n, 0, n)
            s.cell(row=n + 1, column=1).value = n
        rb.save(path)

        # s.write(0, 0, 'step')
        s.cell(row=1, column=1).value = 'step'
        #col = 0
        if Attention_name == 'myc':
            col = 1
        if Attention_name == 'mys':
            col = 2
        if Attention_name == 'my':
            col = 3
        if Attention_name == 'se':
            col = 4
        if Attention_name == 'aa':
            col = 5
        if Attention_name == 'cbam':
            col = 6
        if Attention_name == 'eca':
            col = 7
        if Attention_name == 'none':
            col = 8

        # s.write(0, col, AdaptiveFun_name)
        s.cell(row=1, column=col + 1).value = Attention_name

        i = 1
        while (i <= len(loss)):
            # s.write(i, col, str(loss[i - 1]))  # 像表格中写入数据
            s.cell(row=i + 1, column=col + 1).value = str(loss[i - 1])
            i += 1
        rb.save(path)

def val_loss2excel(step, cross_entropy, Attention_name, dir):
    #path = os.path.join('./Result/ResNet/cifar10_ResNet18_loss_val.xlsx')
    path = os.path.join(dir)
    p = os.path.exists(path)
    loss = cross_entropy
    st = step
    if p:

        rb = openpyxl.load_workbook(path)
        wb = rb.sheetnames[0]
        # wb = copy(rb)
        s = rb[wb]
        # s = wb.get_sheet(0)

        for n in range(1, st + 1):
            # s.write(n, 0, n)
            s.cell(row=n + 1, column=1).value = n
        rb.save(path)

        # s.write(0, 0, 'step')
        s.cell(row=1, column=1).value = 'step'

        if Attention_name == 'myc':
            col = 1
        if Attention_name == 'mys':
            col = 2
        if Attention_name == 'my':
            col = 3
        if Attention_name == 'se':
            col = 4
        if Attention_name == 'aa':
            col = 5
        if Attention_name == 'cbam':
            col = 6
        if Attention_name == 'eca':
            col = 7
        if Attention_name == 'none':
            col = 8

        # s.write(0, col, AdaptiveFun_name)
        s.cell(row=1, column=col + 1).value = Attention_name

        i = 1
        while (i <= len(loss)):
            # s.write(i, col, str(loss[i - 1]))  # 像表格中写入数据
            s.cell(row=i + 1, column=col + 1).value = str(loss[i - 1])
            i += 1
        rb.save(path)


def eval_to_excel(epoch, top1, top3, precision, recall,kappa, Attention_name, dir):
    path = os.path.join(dir)
    #path = os.path.join('./Result/ResNet/cifar10_ResNet18_evolution_train.xls')
    p = os.path.exists(path)

    top1_tr = top1

    top3_tr = top3

    pre = precision

    rec = recall
    ka = kappa
    #maF1_tr = macro_f1

    """
    data = xlwt.Workbook(encoding='ascii')  # 创建一个workboookk
    worksheet = data.add_sheet('Sheet1')  # 添加一个Sheet工作表
    """

    if p:

        rb = xlrd.open_workbook(path)
        wb = copy(rb)
        s = wb.get_sheet(0)

        for n in range(1, epoch + 1):
            s.write(n, 0, n)
        wb.save(path)

        s.write(0, 0, 'epoch')

        if Attention_name == 'myc':
            col = 1
        if Attention_name == 'mys':
            col = 6
        if Attention_name == 'my':
            col = 11
        if Attention_name == 'se':
            col = 16
        if Attention_name == 'aa':
            col = 21
        if Attention_name == 'cbam':
            col = 26
        if Attention_name == 'eca':
            col = 31
        if Attention_name == 'none':
            col = 36

        s.write(0, col, Attention_name + '_top1')
        s.write(0, col + 1, Attention_name + '_top3')
        s.write(0, col + 2, Attention_name + '_pre')
        s.write(0, col + 3, Attention_name + '_rec')
        s.write(0, col + 4, Attention_name + '_F1')



        i = 1
        while (i <= len(top1_tr)):
            s.write(i, col, top1_tr[i - 1])  # 像表格中写入数据
            s.write(i, col + 1, top3_tr[i - 1])  # 像表格中写入数据
            s.write(i, col + 2, pre[i - 1])  # 像表格中写入数据
            s.write(i, col + 3, rec[i - 1])  # 像表格中写入数据
            s.write(i, col + 4, ka[i - 1])  # 像表格中写入数据

            i += 1
        wb.save(path)

